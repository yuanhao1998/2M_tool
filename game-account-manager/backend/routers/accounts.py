from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database import get_db
from models import Account
from schemas import AccountCreate, AccountUpdate, AccountResponse, BatchDelete

router = APIRouter(tags=["accounts"])


@router.get("/accounts", response_model=list[AccountResponse])
def list_accounts(
    search: str = Query(default=""),
    server: str = Query(default=""),
    region: str = Query(default=""),
    email: str = Query(default=""),
    class_name: str = Query(default=""),
    cloud_device: str = Query(default=""),
    db: Session = Depends(get_db),
):
    query = db.query(Account)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            Account.account_name.ilike(pattern) | Account.phone.ilike(pattern)
        )
    if server:
        query = query.filter(Account.server == server)
    if region:
        query = query.filter(Account.region == region)
    if email:
        query = query.filter(Account.email.ilike(f"%{email}%"))
    if class_name:
        query = query.filter(Account.class_name.ilike(f"%{class_name}%"))
    if cloud_device:
        query = query.filter(Account.cloud_device.ilike(f"%{cloud_device}%"))
    return query.order_by(Account.created_at.desc()).all()


@router.get("/accounts/filters")
def list_account_filters(db: Session = Depends(get_db)):
    servers = [
        r[0] for r in db.query(Account.server).filter(Account.server != "").distinct().order_by(Account.server).all()
    ]
    regions = [
        r[0] for r in db.query(Account.region).filter(Account.region != "").distinct().order_by(Account.region).all()
    ]
    return {"servers": servers, "regions": regions}


@router.post("/accounts", response_model=AccountResponse, status_code=201)
def create_account(data: AccountCreate, db: Session = Depends(get_db)):
    if not data.phone or not data.phone.strip():
        raise HTTPException(status_code=422, detail="手机号不能为空")
    if not data.account_name or not data.account_name.strip():
        raise HTTPException(status_code=422, detail="账号名不能为空")
    if not data.cloud_device or not data.cloud_device.strip():
        raise HTTPException(status_code=422, detail="云机名称不能为空")
    existing = db.query(Account).filter(Account.phone == data.phone.strip()).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"手机号 {data.phone} 已存在")
    account = Account(**data.model_dump())
    db.add(account)
    db.commit()
    db.refresh(account)
    return account


HEADER_MAP = {
    "账号名": "account_name", "账号": "account_name",
    "邮箱": "email",
    "区服": "server",
    "大区": "region",
    "职业": "class_name",
    "pin码": "pin_code", "PIN码": "pin_code", "PIN 码": "pin_code",
    "手机号": "phone", "手机": "phone",
    "云机名称": "cloud_device", "云机": "cloud_device",
    "地点": "location",
    "验证码url": "verify_code_url", "验证码URL": "verify_code_url", "验证码 URL": "verify_code_url", "验证码": "verify_code_url",
    "修复码": "recovery_code",
}


@router.post("/accounts/import-excel", response_model=list[AccountResponse], status_code=201)
def import_accounts_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files supported")
    wb = load_workbook(filename=BytesIO(file.file.read()))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Empty file or no data rows")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    field_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = HEADER_MAP.get(h, h)
        field_map[key] = i

    created = []
    updated = []
    for row in rows[1:]:
        data: dict[str, str] = {}
        for field, idx in field_map.items():
            val = str(row[idx]).strip() if row[idx] is not None else ""
            data[field] = val
        phone = data.get("phone", "")
        if not phone:
            continue
        if not data.get("account_name"):
            data["account_name"] = phone

        existing = db.query(Account).filter(Account.phone == phone).first()
        if existing:
            for key, value in data.items():
                setattr(existing, key, value)
            updated.append(existing)
        else:
            account = Account(**data)
            db.add(account)
            created.append(account)

    db.commit()
    for a in created + updated:
        db.refresh(a)
    return created + updated


@router.put("/accounts/{account_id}", response_model=AccountResponse)
def update_account(account_id: int, data: AccountUpdate, db: Session = Depends(get_db)):
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    if data.phone and data.phone != account.phone:
        conflict = db.query(Account).filter(Account.phone == data.phone, Account.id != account_id).first()
        if conflict:
            raise HTTPException(status_code=409, detail=f"手机号 {data.phone} 已被其他账号使用")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(account, key, value)
    db.commit()
    db.refresh(account)
    return account


@router.delete("/accounts/{account_id}", status_code=204)
def delete_account(account_id: int, db: Session = Depends(get_db)):
    account = db.query(Account).filter(Account.id == account_id).first()
    if not account:
        raise HTTPException(status_code=404, detail="Account not found")
    db.delete(account)
    db.commit()
    return None


@router.post("/accounts/batch-delete", status_code=204)
def batch_delete_accounts(data: BatchDelete, db: Session = Depends(get_db)):
    accounts = db.query(Account).filter(Account.id.in_(data.ids)).all()
    for account in accounts:
        db.delete(account)
    db.commit()
    return None
