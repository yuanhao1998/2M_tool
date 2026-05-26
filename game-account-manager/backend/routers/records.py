from datetime import date, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import func
from openpyxl import load_workbook
from io import BytesIO

from database import get_db
from models import Account, DiamondRecord, DiamondSnapshot, DiamondSale
from schemas import DiamondRecordCreate, DiamondRecordBatch, DiamondRecordResponse

router = APIRouter(tags=["records"])


def _sync_snapshots(db: Session, records: list[DiamondRecord]):
    """Create/update snapshots for new records and redistribute gaps."""
    from collections import defaultdict

    accounts_by_date = defaultdict(lambda: defaultdict(int))
    for r in records:
        accounts_by_date[r.account_id][r.recorded_at] += r.amount

    affected_aids = set()
    for aid, date_amounts in accounts_by_date.items():
        # Calculate cumulative diamonds for this account
        all_snaps = (
            db.query(DiamondSnapshot)
            .filter(DiamondSnapshot.account_id == aid)
            .order_by(DiamondSnapshot.date.asc())
            .all()
        )
        cumulative = 0
        date_map = {s.date: s for s in all_snaps}
        if all_snaps:
            cumulative = all_snaps[-1].diamonds

        sorted_dates = sorted(date_amounts.keys())
        for d in sorted_dates:
            cumulative += date_amounts[d]
            existing = date_map.get(d)
            if existing:
                existing.diamonds = cumulative
                existing.change = date_amounts[d]
            else:
                db.add(DiamondSnapshot(
                    account_id=aid, date=d,
                    diamonds=cumulative, change=date_amounts[d],
                ))
                date_map[d] = None  # mark as handled
            affected_aids.add(aid)

    db.flush()

    # Rerun gap distribution for affected accounts
    sales_map = defaultdict(list)
    for s in db.query(DiamondSale).order_by(DiamondSale.id).all():
        sales_map[s.account_id].append((s.sale_date, s.diamonds_sold))

    for aid in affected_aids:
        snaps = (
            db.query(DiamondSnapshot)
            .filter(DiamondSnapshot.account_id == aid)
            .order_by(DiamondSnapshot.date.asc())
            .all()
        )
        if len(snaps) < 2:
            continue

        for i in range(len(snaps) - 1):
            prev = snaps[i]
            curr = snaps[i + 1]
            gap = (curr.date - prev.date).days
            if gap <= 1:
                continue

            total_sold = 0
            for sd in sales_map.get(aid, []):
                if prev.date < sd[0] <= curr.date:
                    total_sold += sd[1]

            if curr.diamonds < prev.diamonds and total_sold > 0:
                total_inc = curr.diamonds + total_sold
            else:
                total_inc = (curr.diamonds - prev.diamonds) + total_sold

            if total_inc <= 0:
                continue

            per_day = total_inc // gap
            extra = total_inc % gap

            d = prev.date + timedelta(days=1)
            while d <= curr.date:
                day_change = per_day + (1 if extra > 0 else 0)
                if extra > 0:
                    extra -= 1

                existing = db.query(DiamondSnapshot).filter(
                    DiamondSnapshot.account_id == aid,
                    DiamondSnapshot.date == d,
                ).first()
                if existing:
                    existing.change = day_change
                else:
                    db.add(DiamondSnapshot(
                        account_id=aid, date=d,
                        diamonds=curr.diamonds if d == curr.date else 0,
                        change=day_change,
                    ))
                d += timedelta(days=1)


@router.get("/records", response_model=list[DiamondRecordResponse])
def list_records(
    date_filter: date | None = Query(default=None, alias="date"),
    account_id: int | None = Query(default=None),
    location: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    query = db.query(DiamondRecord)
    if date_filter:
        query = query.filter(DiamondRecord.recorded_at == date_filter)
    if account_id:
        query = query.filter(DiamondRecord.account_id == account_id)
    if location:
        query = query.filter(DiamondRecord.location.ilike(f"%{location}%"))
    return query.order_by(DiamondRecord.recorded_at.desc(), DiamondRecord.id.desc()).all()


@router.post("/records", response_model=list[DiamondRecordResponse], status_code=201)
def create_records(batch: DiamondRecordBatch, db: Session = Depends(get_db)):
    records = []
    for item in batch.records:
        account = db.query(Account).filter(Account.id == item.account_id).first()
        if not account:
            raise HTTPException(status_code=400, detail=f"Account {item.account_id} not found")
        record = DiamondRecord(
            account_id=item.account_id,
            amount=item.amount,
            location=item.location,
            recorded_at=item.recorded_at,
        )
        db.add(record)
        records.append(record)
    db.commit()
    _sync_snapshots(db, records)
    db.commit()
    for r in records:
        db.refresh(r)
    return records


@router.post("/records/import-excel", response_model=list[DiamondRecordResponse], status_code=201)
def import_excel(
    recorded_at: date = Query(...),
    location: str = Query(default=""),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files supported")
    wb = load_workbook(filename=BytesIO(file.file.read()))
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file or no data rows")

    records = []
    for row in rows:
        if not row[0] or row[1] is None:
            continue
        cloud_device = str(row[0]).strip()
        amount = int(row[1])
        row_location = str(row[2]).strip() if len(row) > 2 and row[2] else location
        account = db.query(Account).filter(Account.cloud_device == cloud_device).first()
        if not account:
            continue
        record = DiamondRecord(
            account_id=account.id,
            amount=amount,
            location=row_location,
            recorded_at=recorded_at,
        )
        db.add(record)
        records.append(record)
    db.commit()
    _sync_snapshots(db, records)
    db.commit()
    for r in records:
        db.refresh(r)
    return records


@router.delete("/records/{record_id}", status_code=204)
def delete_record(record_id: int, db: Session = Depends(get_db)):
    record = db.query(DiamondRecord).filter(DiamondRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    db.delete(record)
    db.commit()
    return None


@router.get("/records/dates")
def list_dates(db: Session = Depends(get_db)):
    dates = (
        db.query(DiamondRecord.recorded_at)
        .distinct()
        .order_by(DiamondRecord.recorded_at.desc())
        .all()
    )
    return [d[0].isoformat() for d in dates]
